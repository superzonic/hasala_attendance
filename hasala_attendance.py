from openpyxl import Workbook, load_workbook
from ATTENDENCE import attendence,summary
import tkinter as Tk
from tkinter import*
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import os




def submit():

    try:
        path_file = open("directory_source_file")
        file_path = str(path_file.read())

        directory_file = open("directory_attendance_file")
        directory_file_path = str(directory_file.read()) + "/attendance_files/"
        if not os.path.exists(directory_file_path):
            os.makedirs(directory_file_path)

        wb = load_workbook(filename=file_path)
        ws = wb.active

        col_num = 2
        name_list = []
        while str(ws['a' + str(col_num)].value) != 'None':
            cell_num = "d" + str(col_num)

            if str(str(ws[cell_num].value)) not in name_list:
                name_list.append(str(str(ws[cell_num].value)))

            col_num += 1

        for iteam in name_list:
            rav = attendence.read(iteam, 7, 6, file_path)
            rav.read_data(directory_file_path)

        summary.summary(str(directory_file_path) + '*.xlsx')

    except :

        messagebox.showerror("showerror", "An error occured. Please check whether you have selected the directory and file correctly")




def working_file():
    #directory = filedialog.askdirectory(title="set the working directory For Invoice MS word files")
    directory = filedialog.askopenfile(title="select the MS excel file for attendance", filetypes=(('MS_EXCEL_FILES', '*.xlsx'),('MS_EXCEL_FILES', '*.xlsx')) )
    path_file = open("directory_source_file", "w")
    path_file.write("")
    path_file = open("directory_source_file", "a")
    path_file.write(directory.name)

def working_directory():
    directory = filedialog.askdirectory(title="set the working directory For Invoice MS word files")
    path_file = open("directory_attendance_file", "w")
    path_file.write("")
    path_file = open("directory_attendance_file", "a")
    path_file.write(directory)






class gui:

    def __init__(self,root):
        self.root = root

    def gui(self):

        self.root = Tk()
        self.root.title('Hasala Attendance')
        self.root.iconbitmap('attendance.ico')
        self.root.geometry('400x100')
        self.root.resizable(False, False)

        toplayerframe = LabelFrame(self.root, padx=0, pady=0)
        toplayerframe.grid(row=0, column=0)

        buttonframe = LabelFrame(self.root, padx=0, pady=0)
        buttonframe.grid(row=1,columnspan=20)

        top_menu = Menu(toplayerframe)
        self.root.config(menu=top_menu)

        file_menu = Menu(top_menu)
        # working directory selection
        top_menu.add_cascade(label="Files", menu=file_menu)
        file_menu.add_command(label="Select the attendance file", command= working_file )
        file_menu.add_separator()
        file_menu.add_command(label="Select the working directory", command= working_directory)

        submit_button = Button(buttonframe, text="Generate", bg="#e9f3e8", font='Helvetica 15 bold', width=25,command=submit)
        submit_button.grid(row=1, columnspan=10,padx=50,pady=15)
        self.root.mainloop()


if __name__ =="__main__":
    gui = gui(root="root")
    gui.gui()
