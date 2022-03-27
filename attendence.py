from openpyxl import Workbook, load_workbook
import os
from datetime import date



class read:

    def __init__(self,name,time_starting_raw,date_starting_raw,filename):
        self.name = name
        self.time_starting_raw=time_starting_raw
        self.date_starting_raw =date_starting_raw
        self.filename = filename

    def read_data(self,directory):

        wb = load_workbook(filename=self.filename)
        ws = wb.active
        i_col_num = 2
        i_cell_num = "i" + str(i_col_num)
        date_time = []
        while str(ws[i_cell_num].value) != 'None':
            i_cell_num = "i" + str(i_col_num)
            d_cell_num = "d" + str(i_col_num)


            if str(str(ws[d_cell_num].value)) == self.name:

                data =str(ws[i_cell_num].value)
                time= data.split(" ")
                date_time.append(time)

            i_col_num += 1


        column_list = ['c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u',
                       'v', 'w', 'x', 'y', 'z', 'aa', 'ab', 'ac', 'ad', 'ae', 'af', 'ag', 'ah', 'ai', 'aj', 'ak', 'al',
                       'am', 'an']

        column_list_iteam = 0

        date = []
        date_iteam = 0

        while date_iteam < len(date_time):
            date.append(date_time[date_iteam][0])

            date_iteam +=1


        date_without_dublicates = []
        [date_without_dublicates.append(x) for x in date if x not in date_without_dublicates]

        date_without_dublicates_modifier = 0

        while date_without_dublicates_modifier < len(date_without_dublicates):
            if date_without_dublicates_modifier == 0:
                if ((date_without_dublicates[0])[8:10]) != "01":
                    date_without_dublicates.insert(0, (date_without_dublicates[date_without_dublicates_modifier])[
                                                      0:8] + "01")
                else:
                    pass
            else:
                while int(((date_without_dublicates[date_without_dublicates_modifier-1])[8:10])) +1 != int(((date_without_dublicates[date_without_dublicates_modifier])[8:10])):
                    if date_without_dublicates_modifier <= 31:
                        date_without_dublicates.insert(date_without_dublicates_modifier,
                                                       (date_without_dublicates[date_without_dublicates_modifier])[
                                                       0:8] + str(int(date_without_dublicates[
                                                                          date_without_dublicates_modifier - 1][
                                                                      8:10]) + 1))
                    else:
                        break


            date_without_dublicates_modifier += 1

        em = load_workbook(filename="em.xlsx")
        es = em.active
        date_without_dublicates_iteam = 0
        while date_without_dublicates_iteam < len(date_without_dublicates):
            es[column_list[column_list_iteam]+str(self.date_starting_raw)]= date_without_dublicates[date_without_dublicates_iteam]
            date_without_dublicates_iteam +=1
            column_list_iteam+=1


        time_counter =0

        while time_counter < len(date_time)-1:
            index_of_col = (date_without_dublicates.index(date_time[time_counter][0]))
            row = self.time_starting_raw

            while True:
                if str(es[column_list[index_of_col]+str(row)].value) == 'None':
                    es[column_list[index_of_col] + str(row)] = date_time[time_counter][1]
                    time_counter += 1
                    break
                else:
                    row += 1

        #this code can be used to insert colums to files. not related to this programe
                #es.insert_cols(5)
                es["a"+str(self.date_starting_raw)] = str(self.name)
                em.save(directory+self.name+".xlsx")

#for testing
if __name__ =="__main__":
        rav = read("RAVNDU      ",5,1,"attendence")
        rav.read_data()